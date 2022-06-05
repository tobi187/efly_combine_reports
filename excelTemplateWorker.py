import dataclasses
import openpyxl as xl
import pandas as pd

DATA_SHEET_NAME = "data"

COLUMN_NAMES = ["Beworbene SKU", "Beworbene ASIN", "Impressionen", "Klicks", "Klickrate (CTR)",
                "Kosten pro Klick (CPC)", "Ausgaben", "7 Tage, Umsatz gesamt (€)",
                "Gesamtumsatzkosten für Werbung (ACOS)",
                "Gesamtrendite von Werbeausgaben (Return on Advertising Spend, ROAS)", "7 Tage, Aufträge gesamt (#)",
                "7Tage, Einheiten gesamt (#)", "7-Tage-Konversionsrate", "7Tage, Beworbene SKU-Einheiten (#)",
                "7-Tage, Andere SKU-Einheiten (#)", "7Tage, Beworbene SKU-Umsätze (€)",
                "7-Tage, Andere SKU-Umsätze (€)", "Gruppierung"]

HEADERS_TO_CHANGE = {
    "Keyword-oderProdukt-Targeting": "Keyword",
    "acos": "Gesamtumsatzkosten für Werbung (ACOS)",
    "Verkäufe": "7 Tage, Aufträge gesamt (#)",
    "Orders": "7 Tage, Umsatz gesamt (€)",
    "14Tage,Einheitengesamt": "7 Tage, Einheiten gesamt (#)",
    "Units": "7 Tage, Einheiten gesamt (#)",
    "Anzeigegruppe": "Anzeigegruppenname",
    "Campaign Name(Informational only)": "Kampagnen-Name",
    "SKU": "Beworbene SKU",
    "ASIN": "Beworbene ASIN",
    "ASIN(Informationalonly)": "Beworbene ASIN",
    "AdGroupName (Informational only)": "Anzeigegruppenname",
    "Portfolio Name": "Portfolioname",
    "Clicks": "Klicks",
    "Click-throughRate": "Klickrate (CTR)",
    "Spend": "Ausgaben",
    "Sales": "7 Tage, Umsatz gesamt (€)",
    "ROAS": "Gesamtrendite von Werbeausgaben (Return on Advertising Spend, ROAS)"
}


class ExcelWorker:
    def __init__(self, fpath: str):
        self.file_path = fpath
        self.row_nr = 2
        self.col_names = {h_name.lower(): index + 1 for index, h_name in enumerate(COLUMN_NAMES)}
        self.double_headers = {h.strip().lower(): i + 1 for i, h in enumerate(HEADERS_TO_CHANGE)}
        # self.unknown_headers = {"placeholder": len(self.col_names) + 1}
        self.unknown_headers = ""

    def change_header(self, df: pd.DataFrame) -> pd.DataFrame:
        for header in df.keys():
            prepped_header = header.replace(" ", "").lower()
            if prepped_header in self.double_headers.keys():
                df.rename({header: self.double_headers[prepped_header]}, axis=1, inplace=True)

    def write_data(self, df: pd.DataFrame) -> None:
        if df.empty:
            return

        df = self.change_header(df)

        wb = self.load_wb()
        sheet = wb[DATA_SHEET_NAME]
        for col_name in df.keys():
            if col_name.strip().lower() in [c.lower() for c in self.col_names.keys()]:
                for index, entry in enumerate(df[col_name]):
                    sheet.cell(row=index + self.row_nr, column=self.col_names[col_name.strip()]).value = entry
            else:
                self.unknown_headers += col_name + ", "
        self.row_nr += len(df.keys()[0])
        wb.save(self.file_path)

    # def unknown_head_index(self):
    #     highest = max(self.unknown_headers.values())
    #     return highest + 1

    def setup(self):
        wb = self.load_wb()
        if DATA_SHEET_NAME not in wb.sheetnames:
            wb.create_sheet(DATA_SHEET_NAME)
        sheet = wb[DATA_SHEET_NAME]

        cols = [sheet.cell(row=1, column=i + 1).value for i in range(sheet.max_row + 1)]
        cols2 = [c.strip() for c in cols if c is not None and c != ""]
        self.col_names = {c: i + 1 for i, c in enumerate(cols2)}

        sheet.delete_rows(1, sheet.max_row + 1)
        for k, v in self.col_names.items():
            sheet.cell(row=1, column=v).value = k
        wb.save(self.file_path)

    def load_wb(self):
        file = self.file_path.split(".")
        if file[-1] == "xlsm":
            return xl.load_workbook(self.file_path, keep_vba=True, read_only=False)
        else:
            return xl.load_workbook(self.file_path)

# old one with spaces
# HEADERS_TO_CHANGE = {
#     "Keyword- oder Produkt-Targeting": "Keyword",
#     "Gesamtumsatz für Werbung (ACoS)": "ACOS ",
#     "Verkäufe": "14 Tage, Umsatz gesamt",
#     "14 Tage, Einheiten gesamt": "Einheiten insgesamt",
#     "Anzeigegruppe": "Anzeigegruppenname",
#     "Campaign Name (Informational only)": "Kampagnen-Name",
#     "Beworbene SKU": "SKU",
#     "Beworbene ASIN": "ASIN",
#     "Ad Group Name (Informational only)": "Anzeigegruppenname",
#     "Portfolioname": "Portfolio Name"
# }
