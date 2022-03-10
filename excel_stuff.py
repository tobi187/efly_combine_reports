import openpyxl as px
import xl_model
import pandas as pd
import xlwings as xw


# def open_ex(path: str) -> list[xl_model.ExToCompare]:
#     file = xw.Book(path)
#     ws: xw.Sheet = file.sheets["Potentielle Kunden DE"]
#
#     data = ws.range((10, 1), (4407, 17)).value
#     df = pd.DataFrame(data)
#     print(df.head(10))
#     print(df.tail(10))

def open_ex(path: str) -> list[xl_model.ExToCompare]:
    wb = px.load_workbook(path, data_only=True)
    ws: px.workbook.workbook.Worksheet = wb["Potentielle Kunden DE"]
    data = []
    # end_of_data = 0
    for i in range(10, 10000009):
        if ws.cell(row=i, column=1).value is None and ws.cell(row=i, column=2).value is None:
            next_ones = [ws.cell(row=r, column=1).value for r in range(i, i+5)]
            if next_ones.count(None) == 5:
                # first empty line
                # end_of_data = i
                break
            continue

        business_data = xl_model.ExToCompare(
            eMail=ws.cell(row=i, column=4).value,
            Unternehmen=ws.cell(row=i, column=1).value,
            Telefon=ws.cell(row=i, column=4).value
            )

        data.append(business_data)

    return data


dt = open_ex(r"C:\Users\fisch\Downloads\CaseStudysTest1.xlsx")
print(dt)
print(len(dt) + 10)
