import openpyxl as xl
import pandas as pd

# start jupiter: jupyter-lab

DATA_SHEET_NAME = "data"
# vscode, pycharm


class ExcelWorker:
    def __init__(self, file_path: str):
        self.start_row = 2
        self.col_names = []
        self.file_path = file_path
        self.double_headers = {k.strip(): v for k, v in change_header.items()}

    def write_data(self, df: pd.DataFrame):
        if df.empty:
            return

        for header in df.keys():
            prepped_header = header.strip()
            if prepped_header in self.double_headers.keys():
                df.rename({header: self.double_headers[prepped_header]}, axis=1, inplace=True)

        col_dic = {}
        for header in df.keys():
            if header in self.col_names:
                col_dic[header] = self.col_names.index(header) + 1
            else:
                col_dic[header] = len(self.col_names) + 1
                self.col_names.append(header)

        wb = self.load_wb()
        sheet = wb[DATA_SHEET_NAME]

        for col_name, col_index in col_dic.items():
            for row_index, entry in enumerate(df[col_name]):
                sheet.cell(row=row_index + self.start_row, column=col_index).value = entry

        for index, col in enumerate(self.col_names):
            sheet.cell(row=1, column=index + 1).value = col

        self.start_row += len(df[df.keys()[0]])
        wb.save(self.file_path)

    def setup(self):
        wb = self.load_wb()
        sheet = wb[DATA_SHEET_NAME]
        sheet.delete_rows(1, sheet.max_row + 1)
        wb.save(self.file_path)

    def load_wb(self):
        file = self.file_path.split(".")
        if file[-1] == "xlsm":
            return xl.load_workbook(self.file_path, keep_vba=True, read_only=False)
        else:
            return xl.load_workbook(self.file_path)


change_header = {
    "Keyword- oder Produkt-Targeting": "Keyword",
    "Gesamtumsatz für Werbung (ACoS)": "ACOS ",
    "Verkäufe": "14 Tage, Umsatz gesamt",
    "14 Tage, Einheiten gesamt": "Einheiten insgesamt",
    "Anzeigegruppe": "Anzeigegruppenname",
    "Campaign Name (Informational only)": "Kampagnen-Name",
    "Beworbene SKU": "SKU",
    "Beworbene ASIN": "ASIN",
    "Ad Group Name (Informational only)": "Anzeigegruppenname",
    "Portfolioname": "Portfolio Name"
}
